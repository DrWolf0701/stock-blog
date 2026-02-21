#!/usr/bin/env python3
"""
PDF 和 Excel 生成功能測試
由 OpenClaw 建立 🐻
"""

import os
import sys
import pandas as pd
from datetime import datetime
from pathlib import Path

# 嘗試導入 PDF 相關套件
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("⚠️ reportlab 未安裝")

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False
    print("⚠️ fpdf 未安裝")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️ openpyxl 未安裝")


def create_sample_data():
    """建立範例資料"""
    print("📊 建立範例資料...")
    
    # 美股新聞範例資料
    news_data = {
        '標題': [
            '美國1月就業報告超預期',
            '台積電ADR勁揚3-4%', 
            'AI投資熱潮持續',
            '企業財報季進行中',
            '資金流向防禦型股票'
        ],
        '分類': ['經濟數據', '個股表現', '產業趨勢', '財報', '資金流向'],
        '重要性': ['高', '高', '中', '中', '中'],
        '影響': ['可能推遲降息', '半導體類股走強', '估值受關注', '關注企業展望', '策略趨保守'],
        '發布時間': ['09:00', '08:45', '10:30', '11:15', '09:45'],
        '來源': ['Bloomberg', 'Reuters', 'CNBC', 'WSJ', 'FT']
    }
    
    df = pd.DataFrame(news_data)
    print(f"✅ 建立 {len(df)} 筆新聞資料")
    return df


def create_excel_file(df, filename="美股新聞.xlsx"):
    """建立 Excel 檔案"""
    if not OPENPYXL_AVAILABLE:
        print("❌ openpyxl 未安裝，無法建立 Excel")
        return False
    
    try:
        print(f"📈 建立 Excel 檔案: {filename}")
        
        # 使用 pandas 建立 Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # 主要資料表
            df.to_excel(writer, sheet_name='新聞摘要', index=False)
            
            # 建立統計表
            stats_df = pd.DataFrame({
                '統計項目': ['新聞數量', '高重要性', '中重要性', '不同來源'],
                '數值': [len(df), len(df[df['重要性'] == '高']), 
                       len(df[df['重要性'] == '中']), df['來源'].nunique()]
            })
            stats_df.to_excel(writer, sheet_name='統計資料', index=False)
            
            # 建立時間分析
            time_analysis = df.groupby('分類').agg({
                '標題': 'count',
                '重要性': lambda x: (x == '高').sum()
            }).rename(columns={'標題': '數量', '重要性': '高重要性數量'})
            time_analysis.to_excel(writer, sheet_name='分類分析')
        
        # 使用 openpyxl 美化格式
        wb = openpyxl.load_workbook(filename)
        
        # 美化新聞摘要表
        ws1 = wb['新聞摘要']
        ws1.column_dimensions['A'].width = 40  # 標題欄位寬度
        ws1.column_dimensions['B'].width = 15
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 25
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 15
        
        # 標題樣式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 重要性顏色標記
        for row in ws1.iter_rows(min_row=2, max_row=len(df)+1, min_col=3, max_col=3):
            for cell in row:
                if cell.value == '高':
                    cell.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                elif cell.value == '中':
                    cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        
        wb.save(filename)
        print(f"✅ Excel 檔案建立完成: {os.path.getsize(filename)} bytes")
        return True
        
    except Exception as e:
        print(f"❌ Excel 建立失敗: {e}")
        return False


def create_pdf_with_reportlab(df, filename="美股新聞.pdf"):
    """使用 reportlab 建立 PDF"""
    if not REPORTLAB_AVAILABLE:
        print("❌ reportlab 未安裝，無法建立 PDF")
        return False
    
    try:
        print(f"📄 建立 PDF 檔案 (reportlab): {filename}")
        
        # 建立 PDF 文件
        doc = SimpleDocTemplate(filename, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # 標題
        title = Paragraph("美股盤前重點新聞彙整", styles['Title'])
        story.append(title)
        
        # 副標題
        subtitle = Paragraph(f"生成時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
        story.append(subtitle)
        story.append(Spacer(1, 12))
        
        # 建立表格資料
        table_data = [['標題', '分類', '重要性', '影響', '時間', '來源']]
        
        for _, row in df.iterrows():
            table_data.append([
                row['標題'],
                row['分類'],
                row['重要性'],
                row['影響'],
                row['發布時間'],
                row['來源']
            ])
        
        # 建立表格
        table = Table(table_data, colWidths=[200, 60, 50, 120, 50, 60])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        # 統計資訊
        stats_text = f"""
        <b>統計摘要：</b><br/>
        總新聞數：{len(df)}<br/>
        高重要性新聞：{len(df[df['重要性'] == '高'])}<br/>
        不同來源數：{df['來源'].nunique()}<br/>
        最新更新時間：{df['發布時間'].max()}
        """
        stats = Paragraph(stats_text, styles['Normal'])
        story.append(stats)
        
        # 生成 PDF
        doc.build(story)
        print(f"✅ PDF 檔案建立完成 (reportlab): {os.path.getsize(filename)} bytes")
        return True
        
    except Exception as e:
        print(f"❌ PDF 建立失敗 (reportlab): {e}")
        return False


def create_pdf_with_fpdf(df, filename="美股新聞_fpdf.pdf"):
    """使用 fpdf 建立 PDF"""
    if not FPDF_AVAILABLE:
        print("❌ fpdf 未安裝，無法建立 PDF")
        return False
    
    try:
        print(f"📄 建立 PDF 檔案 (fpdf): {filename}")
        
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('NotoSansTC', '', '/System/Library/Fonts/PingFang.ttc', uni=True)
        
        # 標題
        pdf.set_font('NotoSansTC', 'B', 16)
        pdf.cell(0, 10, '美股盤前重點新聞彙整', 0, 1, 'C')
        
        pdf.set_font('NotoSansTC', '', 10)
        pdf.cell(0, 10, f'生成時間: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'C')
        pdf.ln(10)
        
        # 表格標題
        pdf.set_font('NotoSansTC', 'B', 10)
        col_widths = [70, 25, 20, 50, 20, 25]
        headers = ['標題', '分類', '重要性', '影響', '時間', '來源']
        
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, 1, 0, 'C')
        pdf.ln()
        
        # 表格內容
        pdf.set_font('NotoSansTC', '', 8)
        for _, row in df.iterrows():
            # 處理換行
            title = row['標題']
            if len(title) > 20:
                title = title[:20] + '...'
            
            pdf.cell(col_widths[0], 10, title, 1)
            pdf.cell(col_widths[1], 10, row['分類'], 1, 0, 'C')
            
            # 重要性顏色標記
            if row['重要性'] == '高':
                pdf.set_text_color(255, 0, 0)  # 紅色
            pdf.cell(col_widths[2], 10, row['重要性'], 1, 0, 'C')
            pdf.set_text_color(0, 0, 0)  # 恢復黑色
            
            pdf.cell(col_widths[3], 10, row['影響'], 1)
            pdf.cell(col_widths[4], 10, row['發布時間'], 1, 0, 'C')
            pdf.cell(col_widths[5], 10, row['來源'], 1, 0, 'C')
            pdf.ln()
        
        # 統計資訊
        pdf.ln(10)
        pdf.set_font('NotoSansTC', 'B', 10)
        pdf.cell(0, 10, '統計摘要:', 0, 1)
        
        pdf.set_font('NotoSansTC', '', 10)
        pdf.cell(0, 8, f'總新聞數: {len(df)}', 0, 1)
        pdf.cell(0, 8, f'高重要性新聞: {len(df[df["重要性"] == "高"])}', 0, 1)
        pdf.cell(0, 8, f'不同來源數: {df["來源"].nunique()}', 0, 1)
        
        pdf.output(filename)
        print(f"✅ PDF 檔案建立完成 (fpdf): {os.path.getsize(filename)} bytes")
        return True
        
    except Exception as e:
        print(f"❌ PDF 建立失敗 (fpdf): {e}")
        return False


def main():
    """主程式"""
    print("=" * 60)
    print("🐻 PDF 和 Excel 生成功能測試")
    print("=" * 60)
    
    # 檢查當前目錄
    current_dir = os.getcwd()
    print(f"工作目錄: {current_dir}")
    
    # 建立範例資料
    df = create_sample_data()
    print(f"\n📋 資料預覽:")
    print(df.to_string())
    
    # 建立 Excel 檔案
    print("\n" + "=" * 40)
    excel_success = create_excel_file(df, "美股新聞測試.xlsx")
    
    # 建立 PDF 檔案 (reportlab)
    print("\n" + "=" * 40)
    pdf_success = create_pdf_with_reportlab(df, "美股新聞_reportlab.pdf")
    
    # 建立 PDF 檔案 (fpdf)
    print("\n" + "=" * 40)
    fpdf_success = create_pdf_with_fpdf(df, "美股新聞_fpdf.pdf")
    
    # 結果總結
    print("\n" + "=" * 60)
    print("📊 測試結果總結:")
    print(f"  Excel 生成: {'✅ 成功' if excel_success else '❌ 失敗'}")
    print(f"  PDF (reportlab): {'✅ 成功' if pdf_success else '❌ 失敗'}")
    print(f"  PDF (fpdf): {'✅ 成功' if fpdf_success else '❌ 失敗'}")
    
    # 列出生成的檔案
    print("\n📁 生成的檔案:")
    for filename in ["美股新聞測試.xlsx", "美股新聞_reportlab.pdf", "美股新聞_fpdf.pdf"]:
        if os.path.exists(filename):
            size = os.path.getsize(filename)
            print(f"  {filename} - {size:,} bytes")
    
    print("\n" + "=" * 60)
    print("🎉 測試完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()